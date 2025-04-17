import pandas as pd
from openpyxl import load_workbook
import argparse

def update_excel_with_dataframe(excel_path, sheet_name, df, start_row=5, start_col=1, column_mapping=None):
    """
    使用DataFrame的值更新Excel文件，同时保持原始格式。
    
    参数:
    excel_path (str): Excel文件的路径
    sheet_name (str): 要更新的工作表名称
    df (DataFrame): 包含更新数据的DataFrame
    start_row (int): 数据开始的行索引（默认为5，因为前4行可能有特殊含义）
    start_col (int): 数据开始的列索引（默认为1，即B列）
    column_mapping (dict): DataFrame列名到Excel列索引的映射（如果不提供，则按顺序映射）
    """
    # 加载现有Excel文件
    print(f"正在加载Excel文件：{excel_path}")
    workbook = load_workbook(excel_path)
    worksheet = workbook[sheet_name]
    
    # 如果没有提供列映射，则创建默认映射（按顺序）
    if column_mapping is None:
        column_mapping = {col_name: i+start_col for i, col_name in enumerate(df.columns)}
    
    # 更新Excel单元格
    print("正在更新Excel单元格...")
    for df_row_idx, row in df.iterrows():
        excel_row_idx = start_row + df_row_idx  # 计算Excel中的行索引
        
        for col_name, value in row.items():
            if col_name in column_mapping:
                excel_col_idx = column_mapping[col_name]
                cell = worksheet.cell(row=excel_row_idx, column=excel_col_idx)
                
                # 保持单元格格式，只更新值
                cell.value = value
    
    # 保存更新后的Excel文件
    print(f"正在保存更新后的Excel文件：{excel_path}")
    workbook.save(excel_path)
    print("Excel文件已成功更新！")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='使用DataFrame更新Excel文件')
    parser.add_argument('--excel', type=str, required=True, help='Excel文件路径')
    parser.add_argument('--sheet', type=str, required=True, help='工作表名称')
    parser.add_argument('--data', type=str, required=True, help='包含更新数据的CSV/Excel文件路径')
    parser.add_argument('--start-row', type=int, default=5, help='数据开始的行索引（默认为5）')
    parser.add_argument('--start-col', type=int, default=1, help='数据开始的列索引（默认为1，即B列）')
    
    args = parser.parse_args()
    
    # 加载更新数据
    if args.data.endswith('.csv'):
        update_df = pd.read_csv(args.data)
    else:
        update_df = pd.read_excel(args.data)
    
    # 执行更新
    update_excel_with_dataframe(
        args.excel, 
        args.sheet, 
        update_df, 
        start_row=args.start_row,
        start_col=args.start_col
    )
